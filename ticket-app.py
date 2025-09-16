import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

filename = "BankClosedTicketReport - 2025-08-28T113909.770.xlsx"
ticket_type = "Consumable"

# Get current date and time
now = datetime.now()
timestamp_str = now.strftime("%Y%m%d_%H%M%S")  # e.g., 20250901_141530
# Dynamic output file name
output_file = f"sla_report_{timestamp_str}.xlsx"

# 1. Read Excel file
df = pd.read_excel(filename, header=16, nrows=15)

# 2. Columns to select
columns_to_print = [
    "TerminalId", "Retailer Name", "TerminalClass", "City",
    "Fault Category", "Within/Outside City",
    "Last Update Date", "Ticket Type", "Last Response Date"
]

# 3. Rename mapping
rename_mapping = {
    "TerminalId": "Terminal ID",
    "Retailer Name": "Merchant Name",
    "TerminalClass": "Standard Merchant",
    "City": "City Name",
    "Last Update Date": "Created On",
    "Within/Outside City": "Within Outside City",
    "Ticket Type": "Call Type Detail",
    "Last Response Date": "Last Response"
}

# 4. Select and rename
selected = df[columns_to_print].rename(columns=rename_mapping).copy()

# 5. Filter by ticket type
selected_df = selected[selected["Call Type Detail"].str.contains(ticket_type, case=False)].copy()

# 6. Convert to datetime
selected_df["Created On"] = pd.to_datetime(selected_df["Created On"]).dt.ceil("s")
selected_df["Last Response"] = pd.to_datetime(selected_df["Last Response"]).dt.ceil("s")

# 7. Split Created On
selected_df["Ticket Open Date"] = selected_df["Created On"].dt.date
selected_df["Ticket Open Time"] = selected_df["Created On"].dt.time
#selected_df["Ticket Open Time Ceil"] = selected_df["Created On"].dt.ceil("S").dt.time

# 8. Split Last Response
selected_df["Last Response Date"] = selected_df["Last Response"].dt.date
selected_df["Last Response Time"] = selected_df["Last Response"].dt.time
#selected_df["Last Response Time Ceil"] = selected_df["Last Response"].dt.ceil("S").dt.time

# 9. Calculate Business-Day Fulfillment (1 business day = 15 hours)
business_hours_per_day = 15

def format_business_timedelta(td, business_hours=15):
    total_seconds = int(td.total_seconds())
    total_hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    seconds = total_seconds % 60

    # Convert total hours into business days
    business_days = total_hours // business_hours
    remaining_hours = total_hours % business_hours

    return f"{business_days} days {remaining_hours:02}:{minutes:02}:{seconds:02}"

# --- Fulfillment Time calculations ---
# 2. Business-day timedelta (15h = 1 day)
selected_df["Fulfillment Time"] = (selected_df["Last Response"] - selected_df["Created On"])\
    .apply(lambda x: format_business_timedelta(x, business_hours_per_day))

# 11. Final column order
final_columns = [
    "Terminal ID", "Merchant Name", "Standard Merchant", "City Name",
    "Within Outside City", "Ticket Open Date", "Ticket Open Time", "Call Type Detail",
    "Last Response Date", "Last Response Time", "Fulfillment Time"
    ]

final_df = selected_df[final_columns]

# 12. Print final report
# print("\n=== Final SLA Report ===\n")
#   print(final_df.to_string(index=False))

# Export to Excel
final_df.to_excel(output_file, index=False)

# Load workbook to adjust column widths
wb = load_workbook(output_file)
ws = wb.active

for col in ws.columns:
    max_length = 0
    column = col[0].column  # Get column number
    column_letter = get_column_letter(column)
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[column_letter].width = max_length + 2  # extra padding

# Save workbook
wb.save(output_file)
print(f"Report exported to '{output_file}' successfully!")


